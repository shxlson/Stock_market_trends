from itertools import count
from matplotlib.animation import FuncAnimation
from matplotlib import pyplot as plt
import openpyxl
import math



#Defining some funcitons for statisctics

#opening Excel sheet and creating an Excel Object
path = "v.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active


#defining Some useful variables and lists
X = []
x = []
Y = []
l = []
R = []
be = 2
e = 334
MA =[]

#Defining some funcitons for statisctics
def sigma_x(x):
    ans = 0
    for elems in x:
        ans += elems
    return ans

def sigma_xy(x,y):
    ans = 0
    for i in range(0,len(x)):
        ans += x[i]*y[i]
    return ans

def sigma_x2(x):
    ans = 0
    for elems in x:
        ans += elems**2
    return ans


for i in range(be,e):
    cell_obj = sheet_obj.cell(row = i, column = 1)
    x.append(str(cell_obj.value))

for i in range(0,len(x)):
    ans = ""
    for j in range(0,9):
        ans += x[i][j]
    x[i] = ans
    ans =""

for i in range(be,e):
    X.append(i)

for i in range(be,e):
    cell_obj = sheet_obj.cell(row = i, column = 2)
    Y.append(cell_obj.value)
Y.reverse()

n = e-be
m = (n*sigma_xy(X,Y) - (sigma_x(X)*sigma_x(Y)))/(n*sigma_x2(X) - sigma_x(X)**2)
b = (sigma_x(Y) -m*sigma_x(X))/n

mas = 5

for i in range(0,n):
    ra = []
    if i<mas:
        MA.append(Y[i])
    else:
        for j in range(i-mas+1,i+1):
            ra.append(Y[j])
        MA.append(sigma_x(ra)/mas)
    ra=[]

for i in range(0,n):
    R.append(m*X[i]+b)

x_data= []
y_data= []
ma_data =[]
r_data = []
fig,ax = plt.subplots()
counter = count(0,1)

def update(i):
    idx=next(counter)
    x_data.append(X[i])
    y_data.append(Y[i])
    r_data.append(R[i])
    ma_data.append(MA[i])
    plt.cla()
    ax.plot(x_data,y_data,"-b",label="Share price")
    ax.plot(x_data,ma_data,"-g",label="Moving averages")
    ax.plot(x_data,r_data,"-r",label="Regression line")
    ax.set_xlabel("Day Number")
    ax.set_ylabel("Share price")  
    ax.legend(loc="upper right")
    
ani = FuncAnimation(fig=fig,func=update,interval=100)
plt.show()
